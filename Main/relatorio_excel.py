import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.chart import PieChart, Reference
from database.repositorio import buscar_gastos_mensais


MESES_PT = {
    1: "Janeiro", 2:"Fevereiro", 3:"Março", 4:"Abril", 5:"Maio", 6:"Junho", 7:"Julho", 8:"Agosto", 9:"Setembro", 10:"Outubro", 11:"Novembro", 12:"Dezembro"
}

def gerar_relatorio_excel(caminho_excel, planilhaFinal, valor_setor):

    gastos = buscar_gastos_mensais()
    df_gastos = pd.DataFrame(gastos, columns=["mes", "ano", "valor total", "registrado em"])

    if not df_gastos.empty:
        df_gastos["mes"] = df_gastos["mes"].map(lambda m: MESES_PT.get(str(m).zfill(2),m))

        df_gastos["registrado em"] = pd.to_datetime(df_gastos["registrado em"]).dt.strftime("%d/%m/%Y %H:%M")

        with pd.ExcelWriter(caminho_excel) as writer:
            planilhaFinal.to_excel(writer, sheet_name="Usuários Válidos", index=False)

            valor_setor.to_excel(writer, sheet_name="Valor-Setor", index= False)
            df_gastos.to_excel(writer, sheet_name="Histórico Mensal", index=False)


    # Estilização Excel
    wb = load_workbook(caminho_excel)

    ws_valor = wb["Valor-Setor"]

    ultima_linha = ws_valor.max_row

    for col in range(1, ws_valor.max_column + 1):
        cell = ws_valor.cell(row=ultima_linha, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
            
    header_fill = PatternFill(start_color="006633", end_color="006633", fill_type="solid")
    linha_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    #Borda
    borda = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"
        
        #Cabeçalho
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = borda

        #Linhas zebra + bordas
        for row in ws.iter_rows(min_row=2):
            if row[0].row % 2==0:
                for cell in row:
                    cell.fill = linha_fill

            for cell in row:
                cell.border = borda
                cell.alignment = Alignment(vertical="center")
                
        # Ajustar largura das colunas
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        ws.auto_filter.ref = ws.dimensions

    wb.save(caminho_excel)